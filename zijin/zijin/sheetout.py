import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime
import os
import mysql.connector
import sys

def main():
    # 从命令行参数读取年级数据
    yesr = sys.argv[1]
    print(yesr)
    # 连接到 MySQL 数据库
    conn = mysql.connector.connect(
        # 正式环境
        # host='192.168.1.7', 
        # port='9908',       
        # user='pxg',     
        # password='Mahiru1206!',
        # database='schooloa', 
        # charset='gbk'
		# 测试环境
		host='154.37.221.97',
		port='8998',
		user='root',
		password='wwwyibu2008',
		database='schooloa',
		charset='gbk'
    )

    today = datetime.now().strftime("%Y-%m-%d")
    now_time = datetime.now().strftime("%H:%M:%S")

    # 判断当前时间是午休还是晚休
    # 用于查询请假时间
    if now_time < "16:00:00":
        e_time = "11:00:00"
        b_time = "12:50:00"
        kb_time = '11:00:00'
    else:
        e_time = "21:00:00"
        b_time = "23:00:00"
        kb_time = "21:30:00"

    # 构建 SQL 查询
    sql = """
    SELECT
        s.`name` AS '姓名',
        s.sex AS '性别',
        s.`status` AS '状态',
        s.out_type AS '住宿类型',
        s.stu_no AS '学号',
        s.yesr AS '年级',
        s.class AS '班级',
        h.hostel_name AS '宿舍',
        r.hostel_room_name AS '房间',
        d.bed_code AS '床位',
        k.kq_date AS '考勤时间',
        k.r_d_Time AS '刷脸时间',
        k.is_in AS '类型', 
        e.STATUS AS '审核状态',
        CONCAT( e.begin_date, ' ', e.begin_time ) AS '请假开始时间',
        CONCAT( e.end_date, ' ', e.end_time ) AS '请假结束时间',
        e.xj_time AS '销假时间',
        t.`name`,
        t.phone
    FROM
        student s
        JOIN hostel_bed d ON d.str_id = s.ID
        JOIN hostel_room r ON r.ID = d.hostel_room_id
        JOIN hostel h ON h.ID = d.hostel_Id
        JOIN room_kq_detail k ON k.stu_no = s.stu_no
        LEFT JOIN stu_leave e ON e.stu_no = s.stu_no 
        AND CONCAT( e.end_date, ' ', e.end_time ) >= '%s %s' 
        AND CONCAT( e.begin_date, ' ', e.begin_time ) <= '%s %s' 
        AND e.`status` IN ( '3', '4' )
        LEFT JOIN teacher_subject b ON b.class_id = s.class_id 
        AND ( b.`status` = '0' )
        LEFT JOIN teacher t ON t.ID = b.Teacher_id 
        AND ( t.`status` = '0' ) 
    WHERE
        s.out_type = '住宿生' 
        AND k.kq_date = CURDATE( ) 
        AND (k.r_d_Time IS NULL or k.r_d_Time <='%s %s')
        AND s.yesr in (%s)
    GROUP BY
        s.stu_no 
    ORDER BY
        h.hostel_name,
        r.hostel_room_name,
        d.bed_code ASC
    """ % (today, e_time, today, b_time, today,kb_time, yesr)

    print("开始导出:" + now_time)
    print("正在导出中.....")
    # 创建游标对象
    cursor = conn.cursor()
    

    # 执行查询
    cursor.execute(sql)

    # 获取所有结果
    rows = cursor.fetchall()

    # 字典
    gaoyi_m = []
    gaoer_m = []
    gaosan_m = []
    gaoyi_w = []
    gaoer_w = []
    gaosan_w = []

    # 序号变量
    en_id = 0
    rn_id = 0
    sn_id = 0
    ew_id = 0
    rw_id = 0
    sw_id = 0

    # 循环存储字典
    for row in rows:
        if row[11] is None and row[13] not in ('3', '4') and row[16] is None:
            if row[5] == "高一":
                if row[1] == "男":
                    en_id += 1
                    gaoyi_m.append([en_id, row[0], row[7], row[8], row[9], row[5], row[6], row[17], row[18], "就寝未归", ""])
                else:
                    ew_id += 1
                    gaoyi_w.append([ew_id, row[0], row[7], row[8], row[9], row[5], row[6], row[17], row[18], "就寝未归", ""])
            elif row[5] == "高二":
                if row[1] == "男":
                    rn_id += 1
                    gaoer_m.append([rn_id, row[0], row[7], row[8], row[9], row[5], row[6], row[17], row[18], "就寝未归", ""])
                else:
                    rw_id += 1
                    gaoer_w.append([rw_id, row[0], row[7], row[8], row[9], row[5], row[6], row[17], row[18], "就寝未归", ""])
            elif row[5] == "高三":
                if row[1] == "男":
                    sn_id += 1
                    gaosan_m.append([sn_id, row[0], row[7], row[8], row[9], row[5], row[6], row[17], row[18], "就寝未归", ""])
                else:
                    sw_id += 1
                    gaosan_w.append([sw_id, row[0], row[7], row[8], row[9], row[5], row[6], row[17], row[18], "就寝未归", ""])

    # 获取当前日期
    current_date = datetime.now().strftime("%Y.%m.%d")
    file_date = datetime.now().strftime("%m.%d")

    # 创建一个新的工作簿
    wb = openpyxl.Workbook()

    # 学校列表
    school_names = [
        {"grade": "男生", "dorm_type": "男生宿舍"},
        {"grade": "女生", "dorm_type": "女生宿舍"},
    ]

    # 定义边框样式
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # 遍历每个宿舍类型
    for idx, dorm_info in enumerate(school_names):
        grade = dorm_info["grade"]
        dorm_type = dorm_info["dorm_type"]

        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()

        time = datetime.now().strftime("%H:%M:%S")
        if time < "16:00:00":
            con = "午休"
        else:
            con = "晚休"

        # 设置表名
        ws.title = f"{dorm_type}"

        # 初始化起始行
        start_row = 1

        if grade == "男生":
            data_groups = [gaoyi_m, gaoer_m, gaosan_m]
            grades = ["高（一）级", "高（二）级", "高（三）级"]
        else:
            data_groups = [gaoyi_w, gaoer_w, gaosan_w]
            grades = ["高（一）级", "高（二）级", "高（三）级"]

        # 处理每个年级的数据
        for data, grade_label in zip(data_groups, grades):
            if data:  # 检查数据列表是否为空
                # 合并单元格并设置年级标题
                title = f"紫金县中山高级中学{grade_label}{grade}就寝未归检查表 {current_date}({con})"
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=11)
                title_cell = ws.cell(row=start_row, column=1)
                title_cell.value = title
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.border = thin_border
                title_cell.font = Font(size=16, bold=True)
                ws.row_dimensions[start_row].height = 15  # 设置标题行高
                start_row += 1

                # 设置表头
                headers = ["序号", "姓名", "宿舍楼", "房间", "床位号", "年级", "班级", "班主任", "班主任电话", "状态", "情况反馈"]
                ws.append(headers)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_num)
                    cell.font = Font(size=11, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws.row_dimensions[start_row].height = 15  # 设置表头行高
                start_row += 1

                # 插入数据
                for row_data in data:
                    ws.append(row_data)
                    last_row = ws.max_row
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=last_row, column=col_num)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        cell.font = Font(size=11)  # 设置数据行字体大小
                    ws.row_dimensions[last_row].height = 15  # 设置数据行高

                # 更新下一个年级的起始行
                start_row = ws.max_row + 1  

        # 调整特定列的宽度，例如班主任电话列
        ws.column_dimensions['I'].width = 15  # 假设班主任电话在第九列，列标为'I'

    # 保存工作簿到文件
    directory = "xlsx"
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_name = f".\\{directory}\\{con}就寝未归考勤明细表({file_date}).xlsx"
    wb.save(file_name)
    print(f"报表已保存为 {file_name}")

    print ("报表导出时间为:" + datetime.now().strftime("%H:%M:%S"))

    # 关闭游标和连接
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()